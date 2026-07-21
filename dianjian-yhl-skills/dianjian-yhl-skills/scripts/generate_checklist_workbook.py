#!/usr/bin/env python3
"""Generate a YHL-style grouped market activity checklist workbook.

Input JSON schema:
{
  "activity": {
    "name": "活动名称",
    "environment": "测试",
    "version_label": "",
    "basics": [
      {"field": "活动时间", "value": "...", "source": "活动规则"}
    ]
  },
  "risks": [
    {"id": "R01", "risk": "...", "finding": "...", "required_confirmation": "..."}
  ],
  "coverage": [
    {"source": "活动规则.docx", "rule": "活动时间", "status": "已覆盖", "target": "A02"}
  ],
  "items": [
    {
      "id": "A01",
      "block": "后端配置确认",
      "group": "活动主配置与发布状态",
      "module": "活动基础",
      "code_entry": "全局",
      "check_item": "后台活动名称",
      "standard": "本次活动口径",
      "action": "后台查看配置",
      "role": "后端配置人员"
    }
  ]
}
"""

from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN = Side(style="thin", color="D9E2EC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
RISK_FILL = PatternFill("solid", fgColor="FCE4D6")
WHITE_BOLD = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
NORMAL = Font(name="Microsoft YaHei", size=10)
TITLE = Font(name="Microsoft YaHei", size=15, bold=True, color="FFFFFF")


BLOCK_ORDER = {"后端配置确认": 0, "前端业务验证": 1}
VALID_ROLES = {"后端配置人员", "前端业务验证"}
LEGACY_ROLES = {"开发", "测试", "产品"}
VALID_ACTIONS = {
    "后台查看配置",
    "页面核对",
    "扫码验证",
    "重复扫码验证",
    "扫码到上限验证",
    "定位/区域验证",
    "核销一张券",
    "1元支付验证",
    "查看后台流水",
    "提现流程验证",
    "导出数据核对",
    "查看确认记录",
}
TIME_KEYWORDS = ("时间", "有效期", "截止", "开始", "结束", "过期")
PHYSICAL_OR_COPY_KEYWORDS = (
    "纸箱",
    "实物",
    "包装",
    "拉环内侧印有",
    "图样",
    "图片",
    "市场总价值",
    "不可兑换现金",
    "个税",
    "热线",
    "合规文案",
)
BROAD_CHECK_ITEMS = {
    "配置预览校验",
    "真实性校验",
    "异常参与风控",
    "凭证核验",
    "合规文案",
    "后台流水",
    "数据看板",
    "数据导出",
    "活动规则展示",
}
NON_CHECK_EXPLANATORY_ITEMS = {
    "配置预览校验",
    "验证说明",
    "测试说明",
    "点检说明",
    "覆盖说明",
    "操作说明",
    "注意事项",
}
NON_CHECK_EXPLANATORY_PHRASES = (
    "能发现",
    "用于",
    "帮助",
    "说明",
    "如何",
    "怎么",
)


def load_input(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def apply_widths(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def setup_print(ws, title, widths):
    ws.sheet_view.showGridLines = False
    apply_widths(ws, widths)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(widths))
    cell = ws.cell(1, 1, title)
    cell.fill = TITLE_FILL
    cell.font = TITLE
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.print_title_rows = "1:3"


def write_header(ws, row_idx, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row_idx, col, header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_idx].height = 22


def write_cell(ws, row_idx, col_idx, value="", center=False):
    cell = ws.cell(row_idx, col_idx, value)
    cell.font = NORMAL
    cell.border = BORDER
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="top",
        wrap_text=True,
    )
    return cell


def sorted_items(items):
    indexed = list(enumerate(items))
    group_order = {}
    for original_idx, item in indexed:
        key = (item.get("block", ""), item.get("group", ""))
        group_order.setdefault(key, original_idx)
    indexed.sort(key=lambda x: (
        BLOCK_ORDER.get(x[1].get("block", ""), 99),
        group_order.get((x[1].get("block", ""), x[1].get("group", "")), x[0]),
        x[0],
    ))
    return [item for _, item in indexed]


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "、".join(stringify(item) for item in value if stringify(item))
    if isinstance(value, dict):
        return "；".join(f"{key}={stringify(val)}" for key, val in value.items())
    return str(value)


def is_time_related_item(item: dict) -> bool:
    # Only treat the row as a time/validity checklist item when its own
    # module, group, or check item is time-related. Standards may mention
    # time as supporting context for unrelated checks, and should not create
    # noisy global-time validation failures.
    text = " ".join(
        stringify(item.get(key, ""))
        for key in ("module", "group", "check_item")
    )
    return any(keyword in text for keyword in TIME_KEYWORDS)


def content_clarity_risk_items(items: list[dict]) -> list[str]:
    risks = []
    for item in items:
        item_id = item.get("id", "")
        check_item = stringify(item.get("check_item", ""))
        standard = stringify(item.get("standard", ""))
        action = stringify(item.get("action", ""))
        text = f"{check_item} {standard}"
        has_risk = False

        if action == "后台查看配置" and any(keyword in text for keyword in PHYSICAL_OR_COPY_KEYWORDS):
            has_risk = True
        if action and action not in VALID_ACTIONS:
            has_risk = True
        separator_count = standard.count("、") + standard.count("；") + standard.count("，")
        if separator_count >= 5 or len(standard) > 85:
            has_risk = True
        if check_item in BROAD_CHECK_ITEMS:
            has_risk = True

        if has_risk and item_id:
            risks.append(item_id)
    return sorted(set(risks))


def non_check_explanatory_items(items: list[dict]) -> list[str]:
    risks = []
    for item in items:
        item_id = item.get("id", "")
        check_item = stringify(item.get("check_item", ""))
        standard = stringify(item.get("standard", ""))
        text = f"{check_item} {standard}"
        has_risk = False

        if check_item in NON_CHECK_EXPLANATORY_ITEMS:
            has_risk = True
        if "错误" in standard and "能发现" in standard:
            has_risk = True
        if any(phrase in check_item for phrase in NON_CHECK_EXPLANATORY_PHRASES) and not (
            "文案" in check_item or "图样说明" in check_item or "规则说明" in check_item
        ):
            has_risk = True
        if check_item in ("说明", "点检说明", "验证说明", "测试说明", "操作说明", "注意事项"):
            has_risk = True

        if has_risk and item_id:
            risks.append(item_id)
    return sorted(set(risks))


def build_activity_sheet(wb, data):
    ws = wb.create_sheet("活动基础口径")
    headers = ["口径项", "本次口径", "来源/备注"]
    write_header(ws, 1, headers)
    activity = data.get("activity", {})
    label_map = {
        "name": "活动名称",
        "environment": "环境",
        "version_label": "版本/构建号",
        "type": "活动类型",
        "time": "活动时间",
        "scan_time": "扫码时间",
        "region": "活动区域",
        "excluded_region": "排除区域",
        "product": "产品范围",
        "code_type": "码类型",
        "reward_type": "奖励类型",
        "merchant_scope": "商户范围",
        "payment_merchant_no": "支付商户号",
        "redeem_merchant_no": "核销商户号",
        "scan_limit": "扫码上限",
        "page_copy_source": "页面文案来源",
    }
    rows = []
    basics = activity.get("basics")
    if isinstance(basics, list):
        for item in basics:
            if isinstance(item, dict):
                rows.append((
                    item.get("field", item.get("name", "")),
                    stringify(item.get("value", item.get("口径", ""))) or "待确认",
                    stringify(item.get("source", item.get("note", ""))),
                ))
    elif isinstance(basics, dict):
        for field, value in basics.items():
            rows.append((field, stringify(value) or "待确认", "activity.basics"))
    for key, label in label_map.items():
        if key == "basics":
            continue
        if key in activity and key not in {"basics"}:
            rows.append((label, stringify(activity.get(key)) or "待确认", "activity"))
    seen = set()
    deduped = []
    for field, value, source in rows:
        key = (field, value)
        if field and key not in seen:
            seen.add(key)
            deduped.append((field, value, source))
    if not deduped:
        deduped = [
            ("活动名称", activity.get("name", "") or "待确认", "activity"),
            ("环境", activity.get("environment", "测试"), "activity"),
            ("版本/构建号", activity.get("version_label", "") or "待确认", "activity"),
        ]
    for row_idx, values in enumerate(deduped, 2):
        for col_idx, value in enumerate(values, 1):
            write_cell(ws, row_idx, col_idx, value)
    apply_widths(ws, [24, 70, 36])
    ws.freeze_panes = "A2"


def build_print_sheet(wb, data):
    activity = data.get("activity", {})
    items = sorted_items(data.get("items", []))
    ws = wb.active
    ws.title = "打印填写版"
    headers = ["编号", "模块", "码入口", "检查项", "检查标准 / 本次口径", "验证动作", "结果", "备注"]
    setup_print(ws, f"{activity.get('name', '活动上线点检表')} - 打印填写版", [9, 13, 14, 30, 58, 30, 9, 30])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    meta = ws.cell(
        2,
        1,
        f"项目名称：{activity.get('name', '')}    环境：{activity.get('environment', '测试')}    验证负责人：__________    验证日期：__________    版本/构建号：__________",
    )
    meta.font = Font(name="Microsoft YaHei", size=11, bold=True)
    meta.border = BORDER
    meta.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20
    write_header(ws, 3, headers)

    row_idx = 4
    current_block = None
    current_group = None
    for item in items:
        block = item.get("block", "未分组复核")
        group = item.get("group", "新增或未归类点检项")
        if block != current_block:
            current_block = block
            current_group = None
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
            cell = ws.cell(row_idx, 1, f"大板块：{block}")
            cell.fill = TITLE_FILL
            cell.font = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1
        if group != current_group:
            current_group = group
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            group_cell = ws.cell(row_idx, 1, f"一次性验证分组：{group}")
            group_cell.fill = SUB_FILL
            group_cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
            group_cell.border = BORDER
            group_cell.alignment = Alignment(vertical="center")
            ws.merge_cells(start_row=row_idx, start_column=6, end_row=row_idx, end_column=8)
            legend = ws.cell(row_idx, 6, "结果填写：√通过 / ×不通过 / -不涉及")
            legend.fill = SUB_FILL
            legend.font = Font(name="Microsoft YaHei", size=11, bold=True)
            legend.border = BORDER
            legend.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

        write_cell(ws, row_idx, 1, item.get("id", ""), True)
        write_cell(ws, row_idx, 2, item.get("module", ""), True)
        write_cell(ws, row_idx, 3, item.get("code_entry", ""), True)
        write_cell(ws, row_idx, 4, item.get("check_item", ""))
        write_cell(ws, row_idx, 5, item.get("standard", ""))
        write_cell(ws, row_idx, 6, item.get("action", ""))
        result = write_cell(ws, row_idx, 7, "", True)
        result.font = Font(name="Microsoft YaHei", size=14, bold=True)
        write_cell(ws, row_idx, 8, "")
        ws.row_dimensions[row_idx].height = 42
        row_idx += 1

    row_idx += 1
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
    footer = ws.cell(row_idx, 1, "整体验收签字：    后端配置人员：__________    前端业务验证：__________    日期：__________")
    footer.fill = SUB_FILL
    footer.font = Font(name="Microsoft YaHei", size=11, bold=True)
    footer.border = BORDER
    ws.print_area = f"A1:H{row_idx}"
    ws.freeze_panes = "A4"


def build_detail_sheet(wb, data):
    ws = wb.create_sheet("上线点检表")
    headers = ["编号", "执行板块", "执行分组", "模块", "码入口", "检查项", "检查标准 / 本次口径", "验证动作", "主责", "环境", "验收结果", "备注"]
    write_header(ws, 1, headers)
    env = data.get("activity", {}).get("environment", "测试")
    for row_idx, item in enumerate(sorted_items(data.get("items", [])), 2):
        values = [
            item.get("id", ""),
            item.get("block", ""),
            item.get("group", ""),
            item.get("module", ""),
            item.get("code_entry", ""),
            item.get("check_item", ""),
            item.get("standard", ""),
            item.get("action", ""),
            item.get("role", ""),
            env,
            "",
            "",
        ]
        for col_idx, value in enumerate(values, 1):
            write_cell(ws, row_idx, col_idx, value, center=col_idx in (1, 2, 9, 10, 11))
    apply_widths(ws, [10, 16, 28, 14, 14, 34, 58, 30, 12, 10, 12, 30])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(1, ws.max_row)}"


def build_risk_sheet(wb, data):
    ws = wb.create_sheet("上线前风险确认")
    headers = ["编号", "风险点", "当前发现", "客户/业务必须确认口径", "确认结果", "备注"]
    write_header(ws, 1, headers)
    for row_idx, risk in enumerate(data.get("risks", []), 2):
        values = [
            risk.get("id", ""),
            risk.get("risk", ""),
            risk.get("finding", ""),
            risk.get("required_confirmation", ""),
            "",
            "",
        ]
        for col_idx, value in enumerate(values, 1):
            write_cell(ws, row_idx, col_idx, value, center=col_idx in (1, 5))
    apply_widths(ws, [10, 28, 56, 56, 14, 34])
    for row in ws.iter_rows(min_row=1, max_row=max(1, ws.max_row)):
        for cell in row:
            if cell.row == 1:
                cell.fill = RISK_FILL
                cell.font = Font(name="Microsoft YaHei", size=10, bold=True)


def build_coverage_sheet(wb, data):
    ws = wb.create_sheet("资料覆盖验证")
    headers = ["资料来源", "规则/需求点", "处理结果", "覆盖编号/风险编号", "说明"]
    write_header(ws, 1, headers)
    coverage = data.get("coverage", data.get("source_coverage", []))
    rows = []
    if isinstance(coverage, list):
        for item in coverage:
            if isinstance(item, dict):
                rows.append([
                    item.get("source", ""),
                    item.get("rule", item.get("requirement", item.get("point", ""))),
                    item.get("status", item.get("result", "")),
                    item.get("target", item.get("item_id", item.get("risk_id", ""))),
                    item.get("note", ""),
                ])
            else:
                rows.append(["", stringify(item), "待人工复核", "", "非结构化覆盖记录"])
    if not rows:
        rows.append(["未提供逐条资料覆盖明细", "生成前需由制表人确认资料已覆盖", "待人工复核", "", ""])
    for row_idx, values in enumerate(rows, 2):
        for col_idx, value in enumerate(values, 1):
            write_cell(ws, row_idx, col_idx, value, center=col_idx in (3, 4))
    apply_widths(ws, [36, 58, 18, 24, 46])
    ws.freeze_panes = "A2"


def build_validation_sheet(wb, data):
    ws = wb.create_sheet("生成校验")
    items = data.get("items", [])
    ids = [item.get("id", "") for item in items]
    counts = Counter(ids)
    duplicates = [code for code, count in counts.items() if code and count > 1]
    missing_ids = [str(idx + 1) for idx, item in enumerate(items) if not item.get("id")]
    ungrouped = [item.get("id", "") for item in items if not item.get("block") or not item.get("group")]
    invalid_blocks = [item.get("id", "") for item in items if item.get("block") not in BLOCK_ORDER]
    global_time_items = [
        item.get("id", "")
        for item in items
        if is_time_related_item(item) and item.get("code_entry", "") in ("", "全局")
    ]
    clarity_risks = content_clarity_risk_items(items)
    non_check_rows = non_check_explanatory_items(items)
    backend_groups = {
        item.get("group", "")
        for item in items
        if item.get("block") == "后端配置确认" and item.get("group")
    }
    frontend_groups = {
        item.get("group", "")
        for item in items
        if item.get("block") == "前端业务验证" and item.get("group")
    }
    legacy_role_items = [
        item.get("id", "")
        for item in items
        if item.get("role", "") in LEGACY_ROLES or item.get("role", "") not in VALID_ROLES
    ]
    legacy_action_items = [
        item.get("id", "")
        for item in items
        if item.get("action", "") and item.get("action", "") not in VALID_ACTIONS
    ]
    legacy_block_items = [
        item.get("id", "")
        for item in items
        if item.get("block", "") not in BLOCK_ORDER
    ]
    coverage = data.get("coverage", data.get("source_coverage", []))
    rows = [
        ("正式点检项数量", len(items)),
        ("唯一编号数量", len(set(ids))),
        ("缺失编号行序号", "、".join(missing_ids) if missing_ids else "无"),
        ("重复编号", "、".join(duplicates) if duplicates else "无"),
        ("未分组编号", "、".join(ungrouped) if ungrouped else "无"),
        ("非标准大板块编号", "、".join(invalid_blocks) if invalid_blocks else "无"),
        ("时间项码入口异常", "、".join(global_time_items) if global_time_items else "无"),
        ("非点检说明性行", "、".join(non_check_rows) if non_check_rows else "无"),
        ("内容清晰度风险", "、".join(clarity_risks) if clarity_risks else "无"),
        ("后端配置确认数量", sum(1 for item in items if item.get("block") == "后端配置确认")),
        ("前端业务验证数量", sum(1 for item in items if item.get("block") == "前端业务验证")),
        ("后端分组数量", len(backend_groups)),
        ("前端业务验证分组数量", len(frontend_groups)),
        ("风险确认项数量", len(data.get("risks", []))),
        ("资料覆盖记录数量", len(coverage) if isinstance(coverage, list) else 0),
        ("旧主责残留编号", "、".join(legacy_role_items) if legacy_role_items else "无"),
        ("旧验证动作残留编号", "、".join(legacy_action_items) if legacy_action_items else "无"),
        ("旧大板块残留编号", "、".join(legacy_block_items) if legacy_block_items else "无"),
        ("打印版结果列", "空白（生成脚本强制为空）"),
        ("打印版备注列", "空白（生成脚本强制为空）"),
    ]
    write_header(ws, 1, ["校验项", "结果"])
    for row_idx, (name, value) in enumerate(rows, 2):
        write_cell(ws, row_idx, 1, name)
        write_cell(ws, row_idx, 2, value)
    apply_widths(ws, [24, 80])


def save_workbook(data, output_path: Path):
    wb = Workbook()
    build_print_sheet(wb, data)
    build_activity_sheet(wb, data)
    build_risk_sheet(wb, data)
    build_detail_sheet(wb, data)
    build_coverage_sheet(wb, data)
    build_validation_sheet(wb, data)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        for row in ws.iter_rows():
            for cell in row:
                cell.border = BORDER
                if cell.value is not None and cell.font == Font():
                    cell.font = NORMAL
    visible_sheets = {"打印填写版"}
    for ws in wb.worksheets:
        ws.sheet_state = "visible" if ws.title in visible_sheets else "hidden"
    wb.active = wb.sheetnames.index("打印填写版")
    wb.save(output_path)


def main(argv):
    if len(argv) != 3:
        print("Usage: generate_checklist_workbook.py input.json output.xlsx", file=sys.stderr)
        return 2
    data = load_input(Path(argv[1]))
    save_workbook(data, Path(argv[2]))
    print(Path(argv[2]).resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
